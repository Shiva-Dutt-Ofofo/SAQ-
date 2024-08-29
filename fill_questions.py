import pandas as pd
from openpyxl import load_workbook

# -- DOC READER - S3 IMPORTS --
import boto3
from io import BytesIO

def find_matching_columns(df, keywords):
    """Find columns that contain any of the specified keywords."""
    matching_columns = []
    for col in df.columns:
        if any(keyword.lower() in col.lower() for keyword in keywords):
            matching_columns.append(col)
    return matching_columns

def find_next_empty_cell(row):
    """Find the index of the next empty cell in a given row."""
    for col_idx, cell in enumerate(row):
        if cell.value is None or cell.value == "":
            return col_idx
    return None  # Return None if no empty cell is found

def extract_answer(string):
    """Extract the answer part of the string after 'Answer:'."""
    if "Answer:" in string:
        return string.split("Answer:")[1].strip()
    return None

def update_answer_column(ws, df, query_column, strings_list):
    """Update the next empty cell in a row based on matches in the query column."""
    query_col_idx = df.columns.get_loc(query_column) + 1  # Adjust for 1-based index in openpyxl
    for idx, query in enumerate(df[query_column]):
        if isinstance(query, str):  # Check if the query is a string
            for string in strings_list:
                if query.strip().lower() in string.lower():
                    next_empty_col_idx = find_next_empty_cell(ws[idx + 2])  # Adjust for header row
                    if next_empty_col_idx is not None:
                        answer = extract_answer(string)  # Extract only the answer part
                        if answer:
                            ws.cell(row=idx + 2, column=next_empty_col_idx + 1, value=answer)  # Write the answer to the next empty cell
                    break  # Move to the next row after finding a match

# def process_excel_file(file_path, strings_list):
#     # Define the keywords to search for in the columns
#     keywords = ["query", "questions", "checklist"]

#     # Load the Excel file with openpyxl
#     wb = load_workbook(file_path)
#     sheet_names = wb.sheetnames

#     # Iterate through each sheet
#     for sheet_name in sheet_names:
#         ws = wb[sheet_name]
#         df = pd.read_excel(file_path, sheet_name=sheet_name)
        
#         # Identify columns that contain the keywords
#         matching_columns = find_matching_columns(df, keywords)

#         # Update the next empty cell in the row for each matching column
#         for column in matching_columns:
#             update_answer_column(ws, df, column, strings_list)

#     # Save the updated workbook back to the Excel file
#     wb.save(file_path)


def process_excel_file(idx, strings_list, bucket_name, prefix):

    # Initialize S3 client
    s3 = boto3.client('s3')

    # # S3 bucket and file information
    # bucket_name = 'your-bucket-name'
    # prefix = 'path/to/your/folder/'  

    # List the files in the S3 bucket
    response = s3.list_objects_v2(Bucket=bucket_name, Prefix=prefix)

    # Filter the files to only include Excel files
    excel_files = [item['Key'] for item in response.get('Contents', []) if item['Key'].endswith('.xlsx')]

    file_key = excel_files[idx]
    print(file_key)
    # Download the file to a BytesIO object
    file_stream = BytesIO()
    s3.download_fileobj(bucket_name, file_key, file_stream)

    # Seek to the beginning of the BytesIO object
    file_stream.seek(0)

    # Define the keywords to search for in the columns
    keywords = ["query", "questions", "checklist"]

    # Load the Excel file with openpyxl
    wb = load_workbook(file_stream)
    sheet_names = wb.sheetnames

    # Iterate through each sheet
    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        df = pd.read_excel(file_stream, sheet_name=sheet_name)
        
        # Identify columns that contain the keywords
        matching_columns = find_matching_columns(df, keywords)

        # Update the next empty cell in the row for each matching column
        for column in matching_columns:
            update_answer_column(ws, df, column, strings_list)

    print('wrote to excel')

    # Save the updated workbook back to the Excel file
    save_to_s3(wb, bucket_name, file_key)


def save_to_s3(workbook, bucket_name, file_key):
    # Save the workbook to a BytesIO object
    output_stream = BytesIO()
    workbook.save(output_stream)
    output_stream.seek(0)

    # object_key = f"{folder_name}/{file_name}"
    file_key_list = file_key.split('/')
    file_key_list[1] = 'result'
    file_key_str = '/'.join(file_key_list)
    print(file_key_str)
    # Upload the file back to S3
    s3 = boto3.client('s3')
    s3.put_object(Bucket=bucket_name, Key=file_key_str, Body=output_stream)
    print('wrote to s3')


# Example usage
# strings_list = [
#     """Question: 1. Describe your application’s architecture and different tiers
# 2. Explain the design and development phases of your application

# Assistant: Answer:

# 1. The Sharpsell application architecture consists of end customers, services, and data management layers. The end customers interact with the application through microservices, which include features like Learning Enablement. External APIs are used for data exchange between the Sharpsell platform and external systems.

# 2. During the design and development phases, the system is designed to meet functional requirements, mitigate risks, and ensure security. The design phase involves identifying potential risks, performing a security risk assessment, and developing a conversion plan. The development phase includes creating detailed documentation, testing, and implementing security vulnerability scans and penetration tests. The operations and maintenance phase involves ongoing system monitoring, modifications, and security vulnerability management.""",
#     """Question: 4. Do you perform web application vulnerability testing? What is the frequency?

# Assistant: Answer: Yes, we perform web application vulnerability testing. The frequency is quarterly.
# """
# ]

qa_pairs = [
    """
    Question: Describe your application’s architecture and different tiers. What are the main components of your application? How does data flow between these components?
    
    Answer:
    The Sharpsell application is a Sales Enablement and Learning platform with a multi-tier architecture. It consists of three main tiers: the presentation tier, the application tier, and the data tier.
    The main components of the application include: third-party APIs for user data and authentication, microservices for each key feature like learning enablement and analytics, and external APIs for data pull from other systems.
    Data flows between these components as follows: end customers interact with the presentation tier, which communicates with the application tier (microservices) for processing requests. The application tier then interacts with the data tier for data storage and retrieval, and third-party APIs for data push and pull.
    """,
    """
    Question: Describe your coding practices.
    
    Answer:
    Our coding practices include unit testing with proper documentation and independent review by a Developer's Manager, integration testing prior to installation in production, and logical access restrictions during testing to ensure no modifications can be made without consent. We also follow secure systems engineering principles and adhere to industry best practices and standards.
    """,
    """
    Question: How do you test your application? What are the testing procedures?
    
    Answer:
    Our application undergoes unit testing, integration testing, and acceptance testing. The Developer supervises unit testing and documents the results on change request forms. The Developer's Manager performs an independent review and signs off on the test results. Integration testing is conducted in a separate environment, and logical access restrictions ensure developers have no update access during testing. Copies of production data or pre-designed test datasets are used for testing purposes. During acceptance testing, logical access restrictions also ensure developers have no update access, and users document problems for developers to address and resubmit for retesting. All significant modifications, major enhancements, and new systems go through these testing procedures before installation in production. IT also monitors vulnerabilities and operating system changes to ensure application control and integrity procedures are not compromised. Access privileges for developers are restricted to prevent unauthorized changes to the code in production and testing environments. Secure SDLC best practices are established, and regular patching and security updates are applied to production and development environments. Annual penetration tests and quarterly vulnerability scans are conducted against internal and external environments. Security awareness training is provided for employees and developers, and an incident response plan is in place for data breaches or security incidents. Risks for key assets and resources are analyzed, and user input is verified and sanitized on both client and server sides. Encrypted channels are used for communications, and unnecessary data is not sent in requests or cookies. Server and application configurations are improved to meet security best practices.
    """,
    """
    Question: Do you perform web application vulnerability testing? What is the frequency?
    
    Answer:
    Yes, we perform web application vulnerability testing. The frequency can vary depending on the client's needs, but it is typically conducted annually or quarterly.
    """,
    """
    Question: If a vulnerability is identified in the software, what is the time required for fixing?
    
    Answer:
    The time required for fixing a vulnerability depends on its severity level. For critical vulnerabilities, the fix is provided as soon as possible within five business days. For high and medium vulnerabilities, the fix is incorporated into the next service pack or the next minor or major release. For low vulnerabilities, the fix is incorporated into the next minor or major release within one month.
    """,
    """
    Question: What is the uptime SLA for the software?
    
    Answer:
    The policy text does not provide information about the uptime SLA for the software.
    """,
    """
    Question: What is the DDOS handling capacity of the application?
    
    Answer:
    I'm unable to determine the DDoS handling capacity of the application from the provided penetration testing report. The report focuses on identifying vulnerabilities and risks related to the Open Web Application Security Project Top 10 Vulnerabilities (OWASP Top 10 - 2021). It does not contain information about the DDoS handling capacity of the application.
    """,
    """
    Question: API keys should be managed and stored securely, public access to the API keys should be prohibited and Key rotation frequency needs to be defined. Could you explain the API key lifecycle management practices mentioned in the document?
    
    Answer:
    The document outlines a 6 month API key rotation schedule, ensuring secure storage within the AWS service, and restricting access solely to authorized personnel. The policy aims to reduce the likelihood of key compromise and maintain strict access controls while minimizing disruption to operations.
    """,
    """
    Question: Deploy use of Single-Sign-On (SSO) between on-premises systems and Cloud environment.
    
    Answer:
    12.2.3 Electronic Messaging: All associates will be informed that Single-Sign-On (SSO) is in use between on-premises systems and the cloud environment. This means they only need to enter their credentials once to access multiple systems. 
    13.1.2 Securing Applications Services on Public Networks: SSO will be implemented as a security measure to protect application services on public networks. By using SSO, the risk of multiple credential sets being compromised is reduced. 
    14.1.1 Information Security Policy for Supplier Relationships: If suppliers are granted access to on-premises systems or cloud environments, SSO will be used to ensure secure access. This is outlined in the Information Security Policy for Supplier Relationships.
    """,
    """
    Question: SaaS application should support the following- SAML, MFA, OTP (not needed if MFA being implemented) & Oauth 2.0. How does the provided document ensure these security features are implemented during the development phase?
    
    Answer:
    The document describes the implementation phase, which includes security vulnerability scans and penetration tests based on the OWASP Top 10 Vulnerabilities. The OWASP Top 10 includes SAML, MFA, and OAuth 2.0 as important security concerns. Therefore, the implementation of these features would be tested and documented during this phase.
    """,
    """
    Question: Do you verify that all of your software/software components suppliers (if any) adhere to industry standards for Systems/Software Development Lifecycle (SDLC) security?
    
    Answer:
    Yes, the Company ensures that all software and software components suppliers comply with industry standards for Systems/Software Development Lifecycle (SDLC) security as detailed in the Company Security Policy and technical control documentation. This includes the identification and documentation of applicable legislation and contractual requirements, as well as the implementation of procedures to ensure compliance with legal restrictions on the use of material in respect of which there may be intellectual property rights. Additionally, all project team members are trained on secure coding standards and the capability of avoiding, finding, and fixing vulnerabilities is planned before project initiation.
    """
]

# file_path = 'Copy of Cloud Checklist_with CSA mapping_v2.4_SaaS (2).xlsx'
result_folder_name = 'prestage-ofofo-storage'
result_file_name = 'res'
bucket_name = 'prestage-ofofo-storage'
#prefix = 'questions/'
prefix = 'questionnaire/query/' + '08a1586d-b8ff-446d-b111-5adc5f0bc877' + '/' + 'f5f0eeaa-18d4-48f2-ab00-0fb37acd0c40' + '/'
idx = 1
process_excel_file(idx, qa_pairs, bucket_name, prefix)


