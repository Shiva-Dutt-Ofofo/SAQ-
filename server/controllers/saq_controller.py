# -- FASTAPI IMPORTS --
# import concurrent.futures
# from fastapi import FastAPI, BackgroundTasks
# import requests
from typing import List

# -- DATA CLEANING & OS OPERATIONS --
# import re
# import os
from dotenv import load_dotenv
from sentence_transformers import SentenceTransformer, util
# import torch

# -- DOC READER - S3 IMPORTS --
import boto3
from io import BytesIO
from openpyxl import load_workbook
import pandas as pd

load_dotenv()

# -- FUNCTION TO RETRIEVE QUESTIONS FROM QUESTION S3 BUCKET --
def get_questions_from_excel(file_details, bucket_name, prefix):
    # Initialize S3 client
    s3 = boto3.client('s3')

    # Extract file name and sheet details from the input parameter
    file_name = file_details["file_name"]
    question_sheet_details = file_details["question_sheet_details"]

    # List the files in the S3 bucket
    response = s3.list_objects_v2(Bucket=bucket_name, Prefix=prefix)

    # Filter the files to find the specific Excel file
    excel_files = [item['Key'] for item in response.get('Contents', []) if item['Key'].endswith('.xlsx')]

    file_key = None
    for key in excel_files:
        if file_name in key:
            file_key = key
            break

    if not file_key:
        raise FileNotFoundError(f"File {file_name} not found in S3 bucket {bucket_name} with prefix {prefix}.")

    print(f"Processing file: {file_key}")

    # Download the file to a BytesIO object
    file_stream = BytesIO()
    s3.download_fileobj(bucket_name, file_key, file_stream)

    # Seek to the beginning of the BytesIO object
    file_stream.seek(0)

    # Load the Excel file with openpyxl
    wb = load_workbook(file_stream)

    # Initialize an empty list to store the extracted questions
    extracted_data = []

    # Iterate through the provided sheet details
    for sheet_detail in question_sheet_details:
        sheet_name = sheet_detail["sheet_name"]
        question_column = sheet_detail["question_column"]
        answer_column = sheet_detail.get("answer_column")

        # Check if the sheet exists in the workbook
        if sheet_name not in wb.sheetnames:
            print(f"Sheet {sheet_name} not found in the workbook. Skipping.")
            continue

        df = pd.read_excel(file_stream, sheet_name=sheet_name)

        # Check if the question column exists in the sheet
        if question_column not in df.columns:
            print(f"Question column {question_column} not found in sheet {sheet_name}. Skipping.")
            continue

        # Check if the answer column exists in the sheet (if provided)
        answer_column_exists = answer_column in df.columns if answer_column else False

        # Extract data from the question column
        for index, question in df[question_column].dropna().iteritems():
            answer = df[answer_column][index] if answer_column_exists else ""
            extracted_data.append({"question": question, "answer": answer})

    print('Extracted data:', extracted_data)
    return extracted_data

# -- FUNCTION THAT UPDATES LIST BASED ON SIMILARITY SEARCH --
def match_and_update_answers(list_one: list, list_two: list, similarity_threshold: float = 0.8):
    # Load the model
    model = SentenceTransformer("all-MiniLM-L6-v2")

    # Extract questions from both lists
    questions_one = [item['question'] for item in list_one]
    questions_two = [item['question'] for item in list_two]

    # Perform paraphrase mining to find similar questions
    paraphrases = util.paraphrase_mining(model, questions_one + questions_two)

    # Create a mapping from questions to their corresponding answers in list_two
    question_to_answer = {item['question']: item['answer'] for item in list_two}

    # Iterate over the found paraphrases and update answers in list_one
    for paraphrase in paraphrases:
        score, i, j = paraphrase

        # Ensure that the similarity is above the threshold
        if score >= similarity_threshold:
            if i < len(list_one) and j >= len(list_one):
                # Match found where i is in list_one and j is in list_two
                list_one[i]['answer'] = question_to_answer[questions_two[j - len(list_one)]]
            elif j < len(list_one) and i >= len(list_one):
                # Match found where j is in list_one and i is in list_two
                list_one[j]['answer'] = question_to_answer[questions_two[i - len(list_one)]]

    return list_one

# -- FUNCTION TO UPDATE ANSWERS TO EXCEL --
def update_excel_and_save_to_s3(file_details, bucket_name, prefix, qa_list):
    # Initialize S3 client
    s3 = boto3.client('s3')

    # Extract file name and sheet details from the input parameter
    file_name = file_details["file_name"]
    question_sheet_details = file_details["question_sheet_details"]

    # List the files in the S3 bucket
    response = s3.list_objects_v2(Bucket=bucket_name, Prefix=prefix)

    # Filter the files to find the specific Excel file
    excel_files = [item['Key'] for item in response.get('Contents', []) if item['Key'].endswith('.xlsx')]

    file_key = None
    for key in excel_files:
        if file_name in key:
            file_key = key
            break

    if not file_key:
        raise FileNotFoundError(f"File {file_name} not found in S3 bucket {bucket_name} with prefix {prefix}.")

    print(f"Processing file: {file_key}")

    # Download the file to a BytesIO object
    file_stream = BytesIO()
    s3.download_fileobj(bucket_name, file_key, file_stream)

    # Seek to the beginning of the BytesIO object
    file_stream.seek(0)

    # Load the Excel file with openpyxl
    wb = load_workbook(file_stream)

    # Iterate through the provided sheet details
    for sheet_detail in question_sheet_details:
        sheet_name = sheet_detail["sheet_name"]
        question_column = sheet_detail["question_column"]
        answer_column = sheet_detail["answer_column"]

        # Check if the sheet exists in the workbook
        if sheet_name not in wb.sheetnames:
            print(f"Sheet {sheet_name} not found in the workbook. Skipping.")
            continue

        ws = wb[sheet_name]
        df = pd.read_excel(file_stream, sheet_name=sheet_name)

        # Check if the question and answer columns exist in the sheet
        if question_column not in df.columns:
            print(f"Question column {question_column} not found in sheet {sheet_name}. Skipping.")
            continue

        if answer_column not in df.columns:
            print(f"Answer column {answer_column} not found in sheet {sheet_name}. Skipping.")
            continue

        # Update the answer column in the Excel sheet for matching questions
        for qa_item in qa_list:
            question = qa_item["question"]
            answer = qa_item["answer"]
            # Find the row where the question matches
            match_index = df[df[question_column] == question].index
            if not match_index.empty:
                # Write the answer to the corresponding cell in the Excel sheet
                ws[f"{answer_column}{match_index[0] + 2}"] = answer

    # Save the updated workbook back to S3
    save_to_s3(wb, bucket_name, file_key)

def save_to_s3(workbook, bucket_name, file_key):
    # Save the workbook to a BytesIO object
    output_stream = BytesIO()
    workbook.save(output_stream)
    output_stream.seek(0)

    # Modify the file key to save the result in a 'result' folder
    file_key_list = file_key.split('/')
    file_key_list[1] = 'result'
    file_key_str = '/'.join(file_key_list)
    print(file_key_str)

    # Upload the file back to S3
    s3 = boto3.client('s3')
    s3.put_object(Bucket=bucket_name, Key=file_key_str, Body=output_stream)
    print('Wrote to S3')

async def security_assesment_questionnaire_result(user_id: str, req_id: str, previous_question_details: List[dict], question_details: List[dict]):

    bucket_name = 'ofofo-stage-storage'
    prefix = 'questionnaire/query/' + user_id + '/' + req_id + '/'

    question_list = get_questions_from_excel(question_details, bucket_name, prefix)

    # prefix = 'questionnaire/previouslyAnsweredQuestionnaire' + user_id + '/' + req_id + '/'
    # question_answer_list = get_questions_from_excel(previous_question_details, bucket_name, prefix)

    # question_list_updated = match_and_update_answers(question_list, question_answer_list)

    # prefix = 'questionnaire/results' + user_id + '/' + req_id + '/'
    # update_excel_and_save_to_s3(question_details, bucket_name, prefix, question_list_updated)

    return {"message": "you"}