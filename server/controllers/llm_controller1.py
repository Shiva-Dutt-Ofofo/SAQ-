# -- IMPORTS --
# -- LANGCHAIN IMPORTS --
from langchain_text_splitters import RecursiveCharacterTextSplitter
from langchain.chains import RetrievalQA
from langchain_community.llms import HuggingFaceEndpoint
from langchain_huggingface import HuggingFaceEmbeddings
from langchain_pinecone import PineconeVectorStore
from langchain_community.document_loaders import S3DirectoryLoader
from langchain.chains import LLMChain
from langchain_core.prompts import PromptTemplate
from pinecone import Pinecone
from langchain.chains import create_retrieval_chain
from langchain.chains.combine_documents import create_stuff_documents_chain
from langchain_core.prompts import ChatPromptTemplate
import concurrent.futures
from fastapi import FastAPI, BackgroundTasks
import requests

# -- DOC READER - S3 IMPORTS --
import boto3
from io import BytesIO
from openpyxl import load_workbook
import pandas as pd
from openpyxl.utils import column_index_from_string
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import MergedCell

# -- DATA CLEANING & OS OPERATIONS --
import re
import os
from dotenv import load_dotenv
from sentence_transformers import SentenceTransformer, util
import torch
from pinecone import Pinecone, ServerlessSpec

#import nltk
#nltk.download('wordnet')
load_dotenv()

# -- NEWER CREATE RETRIEVAL CHAIN SETUP --
def new_retrieval_chain(vector_database, input_query: str):

    repo_id = "meta-llama/Meta-Llama-3.1-8B-Instruct"
    # repo_id = "mistralai/Mistral-7B-Instruct-v0.2"

    large_language_model = HuggingFaceEndpoint(
        repo_id=repo_id, temperature=0.3, token=os.environ['HUGGINGFACEHUB_API_TOKEN']
    )

    system_prompt = (
        "You are an assistant that does not deviate from the instructions for answering a question. "
        "Use the following pieces of retrieved context to answer "
        "the question mentioned in the human prompt. If you don't know the answer, say that you "
        "don't know. Use three sentences maximum and keep the "
        "answer concise."
        "\n\n"
        "It is very important that you only answer to the question given by the human and you should never make up your own question"
        "The structure of your answer should always be:"
        "Answer: answer to the question"
        "{context}"
    )

    prompt = ChatPromptTemplate.from_messages(
        [
            ("system", system_prompt),
            ("human", "{input}"),
        ]
    )

    retriever = vector_database.as_retriever()

    question_answer_chain = create_stuff_documents_chain(large_language_model, prompt)
    rag_chain = create_retrieval_chain(retriever, question_answer_chain)

    response = rag_chain.invoke({"input": input_query})
    return response["answer"]
    # print(response["context"])

# -- LOAD DOCUMENTS FROM S3 BUCKET --
def load_documents_from_s3_directory(directory_name, prefix):

    """Loads files (all types) from the specified S3 directory.

    Args:
        directory_name: The name of the S3 directory.
        prefix: Prefix of the specified S3 directory.

    Response:
        List of Loaded documents from the S3 directory.
    """


    document_loader = S3DirectoryLoader(
        directory_name, prefix = prefix)

    response = document_loader.load()
    print('response ::::',response)
    return response

# -- TEXT SPLITTER HELPER FUNCTION --
def chunk_data(documents, chunk_size, chunk_overlap):

    """ Helper function that splits the incoming documents into chunks for further processing.
    
    Args: 
        documents (List): Raw documents loaded from the document loader.
        chunk_size (int): Size of a particular chunk for splitting 
        chunk_overlap (int): Overlap to maintain context between each chunk.
    """
    # print(documents)
    documents[0].page_content = re.sub(r'\s{3,}', ' ', documents[0].page_content)

    text_splitter = RecursiveCharacterTextSplitter(chunk_size=chunk_size,chunk_overlap=chunk_overlap)
    documents = text_splitter.split_documents(documents)

    return documents

# -- FUNCTION THAT CREATES VECTOR STORE BASED ON INDEX NAME AND DOCUMENTS --
def create_vector_store_from_docs(documents, index_name):
   """Creates a vector store by processing documents with multithreading and adding chunks directly to the vector store."""
   Pinecone(api_key=os.environ['PINECONE_API_KEY'])
   model_name = "sentence-transformers/all-mpnet-base-v2"
   model_kwargs = {'device': 'cpu'}
   encode_kwargs = {'normalize_embeddings': False}
   embeddings = HuggingFaceEmbeddings(
       model_name=model_name,
       model_kwargs=model_kwargs,
       encode_kwargs=encode_kwargs
   )
   # Initialize the vector store
   vectorstore = PineconeVectorStore(index_name=index_name, embedding=embeddings)
   # Split documents into 4 parts for parallel processing
   num_docs = len(documents)
   split_docs = [documents[i::4] for i in range(4)]
   def worker(docs):
       """Worker function to process a chunk of documents and add to the vector store."""
       chunks = chunk_data(docs, 3500, 100)
       vectorstore.add_documents(chunks)
   # Execute the worker function in parallel using ThreadPoolExecutor
   with concurrent.futures.ThreadPoolExecutor(max_workers=4) as executor:
       executor.map(worker, split_docs)
   return vectorstore

# # -- MULTITHREADED FUNCTION THAT INVOKES LLM --
# def fetch_result_from_context(question_list, vector_store):

#     """For all the questions present in the document, fetch answers from the context and append to a list.
 
#     Args:
#         question_list (list): A list of questions.
#         vector_store: The vector store.
 
#     Returns:
#         tuple: A tuple containing a big string with questions followed by answers and a list of individual question-answer pairs.
#     """

#     max_workers = 4  # Adjust this value based on your system and workload

#     def process_questions(questions_chunk):
#         results = []
#         for question in questions_chunk:
#             query = f"{question}"
#             result = f"Question: {question}\n"
#             result += new_retrieval_chain(vector_store, query)
#             result += '\n\n'
#             results.append(result)
#         return results

#     # Determine the size of each chunk for multithreading
#     chunk_size = len(question_list) // max_workers
#     if chunk_size == 0:
#         chunk_size = 1
    
#     chunks = [question_list[i:i + chunk_size] for i in range(0, len(question_list), chunk_size)]

#     combined_results = []
#     results_list = []

#     # Use ThreadPoolExecutor to process questions in parallel
#     with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
#         futures = [executor.submit(process_questions, chunk) for chunk in chunks]
#         for future in concurrent.futures.as_completed(futures):
#             chunk_results = future.result()
#             results_list.extend(chunk_results)
#             combined_results.append(''.join(chunk_results))

#     # Combine all results into a single string
#     combined_results_string = ''.join(combined_results)

#     return combined_results_string, results_list

# -- SINGLETHREADED FUNCTION THAT INVOKES LLM RAG - RETRIEVAL CHAIN -- 
def fetch_result_from_context(questions_list, vector_store):
    """For all the questions present in the document, fetch answers from the context and append to the respective question's answer field.
    
    Args:
        questions_list (list): A list of dictionaries where each dictionary has 'question' and 'answer' fields
        vector_store (object): The vector store or database to retrieve the answers from
    
    Returns:
        list: The updated questions_list with the 'answer' field filled in for each question
    """
    
    for item in questions_list:
        question = item['question']
        
        query = f"{question}"
        
        # Fetch the answer from the retrieval chain using the vector store
        answer = new_retrieval_chain(vector_store, query)
        
        # Update the answer field of the current question
        item['answer'] = answer
    
    return questions_list

# -- FUNCTION TO RETRIEVE QUESTIONS FROM QUESTION S3 BUCKET --
def get_questions_from_excel(file_details_list, bucket_name, prefix):
    # Initialize S3 client
    s3 = boto3.client('s3')

    # Initialize an empty list to store the extracted questions from all files
    all_extracted_data = []

    # Iterate over each file's details in the list
    for file_details in file_details_list:
        file_name = file_details["file_name"]
        question_sheet_details = file_details["question_sheet_details"]

        # List the files in the S3 bucket
        response = s3.list_objects_v2(Bucket=bucket_name, Prefix=prefix)

        # Filter the files to find the specific Excel file
        excel_files = [item['Key'] for item in response.get('Contents', []) if item['Key'].endswith('.xlsx')]

        file_key = None
        for key in excel_files:
            if file_name in key:
                file_key = key
                break

        if not file_key:
            raise FileNotFoundError(f"File {file_name} not found in S3 bucket {bucket_name} with prefix {prefix}.")

        print(f"Processing file: {file_key}")

        # Download the file to a BytesIO object
        file_stream = BytesIO()
        s3.download_fileobj(bucket_name, file_key, file_stream)

        # Seek to the beginning of the BytesIO object
        file_stream.seek(0)

        # Load the Excel file with openpyxl
        wb = load_workbook(file_stream)

        # Iterate through the provided sheet details
        for sheet_detail in question_sheet_details:
            sheet_name = sheet_detail["sheet_name"]
            question_column = sheet_detail["question_column"]
            answer_column = sheet_detail.get("answer_column")

            # Check if the sheet exists in the workbook
            if sheet_name not in wb.sheetnames:
                print(f"Sheet {sheet_name} not found in the workbook. Skipping.")
                continue

            ws = wb[sheet_name]

            # Convert column letters to indices
            question_col_idx = column_index_from_string(question_column) - 1
            answer_col_idx = column_index_from_string(answer_column) - 1 if answer_column else None

            # Iterate over the rows of the sheet and extract questions and answers
            for row in ws.iter_rows(min_row=2, values_only=True):
                question = row[question_col_idx]
                answer = row[answer_col_idx] if answer_col_idx is not None else ""

                if question:
                    all_extracted_data.append({"question": question, "answer": answer})

    print('Extracted data:', all_extracted_data)
    return all_extracted_data


# -- FUNCTION TO UPDATE ANSWERS TO EXCEL --
def update_excel_and_save_to_s3(file_details_list, bucket_name, prefix, qa_list):
    # Initialize S3 client
    s3 = boto3.client('s3')

    # Iterate over each file's details in the list
    for file_details in file_details_list:
        file_name = file_details["file_name"]
        question_sheet_details = file_details["question_sheet_details"]

        # List the files in the S3 bucket
        response = s3.list_objects_v2(Bucket=bucket_name, Prefix=prefix)

        # Filter the files to find the specific Excel file
        excel_files = [item['Key'] for item in response.get('Contents', []) if item['Key'].endswith('.xlsx')]

        file_key = None
        for key in excel_files:
            if file_name in key:
                file_key = key
                break

        if not file_key:
            raise FileNotFoundError(f"File {file_name} not found in S3 bucket {bucket_name} with prefix {prefix}.")

        print(f"Processing file: {file_key}")

        # Download the file to a BytesIO object
        file_stream = BytesIO()
        s3.download_fileobj(bucket_name, file_key, file_stream)

        # Seek to the beginning of the BytesIO object
        file_stream.seek(0)

        # Load the Excel file with openpyxl
        wb = load_workbook(file_stream)

        # Iterate through the provided sheet details
        for sheet_detail in question_sheet_details:
            sheet_name = sheet_detail["sheet_name"]
            question_column = sheet_detail["question_column"]
            answer_column = sheet_detail["answer_column"]

            # Check if the sheet exists in the workbook
            if sheet_name not in wb.sheetnames:
                print(f"Sheet {sheet_name} not found in the workbook. Skipping.")
                continue

            ws = wb[sheet_name]

            # Convert column letters to indices
            question_col_idx = column_index_from_string(question_column) - 1
            answer_col_idx = column_index_from_string(answer_column) - 1

            # Load data into pandas dataframe for easier row matching
            df = pd.DataFrame(ws.iter_rows(values_only=True))

            # Update the answer column in the Excel sheet for matching questions
            for qa_item in qa_list:
                question = qa_item["question"]
                answer = qa_item["answer"]

                # Find the row where the question matches
                match_index = df[df[question_col_idx] == question].index
                if not match_index.empty:
                    row_idx = match_index[0] + 2
                    col_idx = answer_col_idx + 1

                    # Detect if the current cell is part of a merged range
                    for merged_cell_range in ws.merged_cells.ranges:
                        if (row_idx >= merged_cell_range.min_row and row_idx <= merged_cell_range.max_row) and \
                           (col_idx >= merged_cell_range.min_col and col_idx <= merged_cell_range.max_col):
                            print(f"Writing to the top-left of the merged cell at: {(merged_cell_range.min_row, merged_cell_range.min_col)}")
                            ws.cell(row=merged_cell_range.min_row, column=merged_cell_range.min_col, value=answer)
                            break
                    else:
                        # If not a merged cell, write the answer to the corresponding cell
                        ws.cell(row=row_idx, column=col_idx, value=answer)

        # Save the updated workbook back to S3
        save_to_s3(wb, bucket_name, file_key)




def save_to_s3(workbook, bucket_name, file_key):
    # Save the workbook to a BytesIO object
    output_stream = BytesIO()
    workbook.save(output_stream)
    output_stream.seek(0)

    # Modify the file key to save the result in a 'result' folder
    file_key_list = file_key.split('/')
    file_key_list[1] = 'result'
    file_key_str = '/'.join(file_key_list)
    print(file_key_str)

    # Upload the file back to S3
    s3 = boto3.client('s3')
    s3.put_object(Bucket=bucket_name, Key=file_key_str, Body=output_stream)
    print('Wrote to S3')

async def retrieval_llama(user_id: str, req_id: str, question_details: list[dict], index_present: bool):



    if(index_present):
        # -- CREATING VECTOR STORE --
        print('call started')
        bucket_name = 'prestage-ofofo-storage'
        prefix = 'questionnaire/context/' + user_id + '/' + req_id + '/'
        context_file_data = load_documents_from_s3_directory(bucket_name, prefix)
        print("reading context done..!")

        # # Split the documents into chunks and create a vector store from the chunks
        documents = chunk_data(context_file_data, 3500, 100)
        # index_name = user_id
        index_name = 'docs-rag-chatbot'
        create_vector_store_from_docs(documents, index_name)
    else:
        #create vector store
        print('create index')
        pc = Pinecone(api_key="9a3ec25c-b9dc-4656-9555-5fecbf42bf8c")
        index_name = user_id
        pc.create_index(
            name=index_name,
            dimension=768, # Replace with your model dimensions
            metric="cosine", # Replace with your model metric
            spec=ServerlessSpec(
                cloud="aws",
                region="us-east-1"
            ) 
        )
        print('call started')
        bucket_name = 'prestage-ofofo-storage'
        prefix = 'questionnaire/context/' + user_id + '/' + req_id + '/'
        context_file_data = load_documents_from_s3_directory(bucket_name, prefix)
        print("reading context done..!")

        # # Split the documents into chunks and create a vector store from the chunks
        documents = chunk_data(context_file_data, 3500, 100)
        create_vector_store_from_docs(documents, index_name)
        

    embedding_function = HuggingFaceEmbeddings()

    # index_name = user_id
    index_name = 'docs-rag-chatbot'
    vector_store = PineconeVectorStore(
        index_name = index_name, embedding=embedding_function
    )

    bucket_name = 'prestage-ofofo-storage'
    prefix = 'questionnaire/result/' + user_id + '/' + req_id + '/' # fetch from results folder later

    question_list = get_questions_from_excel(question_details, bucket_name, prefix)
    filtered_list = [item for item in question_list if item['answer'] is None]
    print("Read Questions :::")
    updated_question_list = fetch_result_from_context(filtered_list, vector_store)
    print("Questions Generated::::")

    prefix = 'questionnaire/result/' + user_id + '/' + req_id + '/'
    update_excel_and_save_to_s3(question_details, bucket_name, prefix, updated_question_list)

    

