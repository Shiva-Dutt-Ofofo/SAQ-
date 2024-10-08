# -- FASTAPI IMPORTS --
# import concurrent.futures
# from fastapi import FastAPI, BackgroundTasks
# import requests
from typing import List

# -- DATA CLEANING & OS OPERATIONS --
# import re
# import os
from dotenv import load_dotenv
from sentence_transformers import SentenceTransformer, util
# import torch

# -- DOC READER - S3 IMPORTS --
import boto3
from io import BytesIO
from openpyxl import load_workbook
import pandas as pd
from openpyxl.utils import column_index_from_string
import spacy

load_dotenv()


# -- FUNCTION TO RETRIEVE QUESTIONS FROM QUESTION S3 BUCKET --
def get_questions_from_excel(file_details, bucket_name, prefix, previous_questionnaire: bool):
    # Initialize S3 client
    s3 = boto3.client('s3')

    # Extract file name and sheet details from the input parameter
    file_name = file_details["file_name"]
    question_sheet_details = []
    if(previous_questionnaire):
        question_sheet_details = file_details["previous_question_sheet_details"]
    else:
        question_sheet_details = file_details["question_sheet_details"]

        
    # List the files in the S3 bucket
    response = s3.list_objects_v2(Bucket=bucket_name, Prefix=prefix)

    # Filter the files to find the specific Excel file
    excel_files = [item['Key'] for item in response.get('Contents', []) if item['Key'].endswith('.xlsx')]

    file_key = None
    for key in excel_files:
        if file_name in key:
            file_key = key
            break

    if not file_key:
        raise FileNotFoundError(f"File {file_name} not found in S3 bucket {bucket_name} with prefix {prefix}.")

    print(f"Processing file: {file_key}")

    # Download the file to a BytesIO object
    file_stream = BytesIO()
    s3.download_fileobj(bucket_name, file_key, file_stream)

    # Seek to the beginning of the BytesIO object
    file_stream.seek(0)

    # Load the Excel file with openpyxl
    wb = load_workbook(file_stream)

    # Initialize an empty list to store the extracted questions
    extracted_data = []

    # Iterate through the provided sheet details
    for sheet_detail in question_sheet_details:
        sheet_name = sheet_detail["sheet_name"]
        question_column = sheet_detail["question_column"]  # This is a letter like "A", "B", etc.
        answer_column = sheet_detail.get("answer_column")  # This is also a letter like "B", "C", etc.

        # Check if the sheet exists in the workbook
        if sheet_name not in wb.sheetnames:
            print(f"Sheet {sheet_name} not found in the workbook. Skipping.")
            continue

        ws = wb[sheet_name]

        # Convert the column letters to indices (1-based, so subtract 1 for 0-based indexing)
        question_col_idx = column_index_from_string(question_column) - 1
        answer_col_idx = column_index_from_string(answer_column) - 1 if answer_column else None

        # Iterate over the rows of the sheet and extract questions and answers
        for row in ws.iter_rows(min_row=2, values_only=True):  # Start from row 2 to skip headers
            question = row[question_col_idx]
            answer = row[answer_col_idx] if answer_col_idx is not None else ""

            if question:  # Ensure that question is not empty
                extracted_data.append({"question": question, "answer": answer})

    print('Extracted data:', extracted_data)
    return extracted_data

# -- FUNCTION THAT UPDATES LIST BASED ON SIMILARITY SEARCH --
def match_and_update_answers(list_one: list, list_two: list, similarity_threshold: float = 0.7):
    # Load the spaCy model with word vectors
    nlp = spacy.load("en_core_web_md")  # or "en_core_web_lg" for larger models

    # Create a mapping from questions to their corresponding answers in list_two
    question_to_answer = {item['question']: item['answer'] for item in list_two}

    # Iterate over each question in list_one
    for i, item_one in enumerate(list_one):
        max_similarity = 0
        best_match_answer = None
        
        # Convert the current question in list_one to a spaCy Doc object
        question_one = nlp(item_one['question'])

        # Compare it with every question in list_two
        for item_two in list_two:
            # Convert the current question in list_two to a spaCy Doc object
            question_two = nlp(item_two['question'])

            # Compute the similarity score
            similarity_score = question_one.similarity(question_two)

            # If the similarity score is above the threshold and higher than the previous best match
            if similarity_score >= similarity_threshold and similarity_score > max_similarity:
                max_similarity = similarity_score
                best_match_answer = item_two['answer']

        # If a similar question was found, update the answer in list_one
        if best_match_answer:
            list_one[i]['answer'] = best_match_answer

    return list_one


def update_excel_and_save_to_s3(file_details, bucket_name, prefix, qa_list):
    # Initialize S3 client
    s3 = boto3.client('s3')

    # Extract file name and sheet details from the input parameter
    file_name = file_details["file_name"]
    question_sheet_details = file_details["question_sheet_details"]

    # List the files in the S3 bucket
    response = s3.list_objects_v2(Bucket=bucket_name, Prefix=prefix)

    # Filter the files to find the specific Excel file
    excel_files = [item['Key'] for item in response.get('Contents', []) if item['Key'].endswith('.xlsx')]

    file_key = None
    for key in excel_files:
        if file_name in key:
            file_key = key
            break

    if not file_key:
        raise FileNotFoundError(f"File {file_name} not found in S3 bucket {bucket_name} with prefix {prefix}.")

    print(f"Processing file: {file_key}")

    # Download the file to a BytesIO object
    file_stream = BytesIO()
    s3.download_fileobj(bucket_name, file_key, file_stream)

    # Seek to the beginning of the BytesIO object
    file_stream.seek(0)

    # Load the Excel file with openpyxl
    wb = load_workbook(file_stream)

    # Iterate through the provided sheet details
    for sheet_detail in question_sheet_details:
        sheet_name = sheet_detail["sheet_name"]
        question_column = sheet_detail["question_column"]  # Position like "A", "B", etc.
        answer_column = sheet_detail["answer_column"]  # Position like "B", "C", etc.

        # Check if the sheet exists in the workbook
        if sheet_name not in wb.sheetnames:
            print(f"Sheet {sheet_name} not found in the workbook. Skipping.")
            continue

        ws = wb[sheet_name]

        # Convert column letters to indices (1-based, subtract 1 for 0-based indexing)
        question_col_idx = column_index_from_string(question_column)
        answer_col_idx = column_index_from_string(answer_column)

        # Iterate through rows to find the matching question and update the answer
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):  # Skipping header row
            row_question = row[question_col_idx - 1].value  # 1-based index to 0-based
            for qa_item in qa_list:
                if row_question == qa_item["question"]:
                    # Update the answer in the corresponding answer column
                    ws.cell(row=row[0].row, column=answer_col_idx, value=qa_item["answer"])
                    break

    # Save the updated workbook back to S3
    save_to_s3(wb, bucket_name, file_key)
    
def save_to_s3(workbook, bucket_name, file_key):
    # Save the workbook to a BytesIO object
    output_stream = BytesIO()
    workbook.save(output_stream)
    output_stream.seek(0)

    # Modify the file key to save the result in a 'result' folder
    file_key_list = file_key.split('/')
    file_key_list[1] = 'result'
    file_key_str = '/'.join(file_key_list)
    print(file_key_str)

    # Upload the file back to S3
    s3 = boto3.client('s3')
    s3.put_object(Bucket=bucket_name, Key=file_key_str, Body=output_stream)
    print('Wrote to S3')

async def security_assesment_questionnaire_result(user_id: str, req_id: str, previous_question_details: List[dict], question_details: List[dict]):
    print('start')
    bucket_name = 'prestage-ofofo-storage'
    prefix = 'questionnaire/query/' + user_id + '/' + req_id + '/'
    print('start')
    question_list = get_questions_from_excel(question_details, bucket_name, prefix, False)
    print('start')
    prefix = 'questionnaire/previouslyAnsweredQuestionnaire/' + user_id + '/' + req_id + '/'
    question_answer_list = get_questions_from_excel(previous_question_details, bucket_name, prefix, True)
    print('start')
    question_list_updated = match_and_update_answers(question_list, question_answer_list)
    print('UPDATED ::::::')
    print(question_list_updated)
    prefix = 'questionnaire/query/' + user_id + '/' + req_id + '/'
    update_excel_and_save_to_s3(question_details, bucket_name, prefix, question_list_updated)
    print('start')
    return {"message": "you"}